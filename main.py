#필요한 모듈 불러오기
#만약 여기서 에러가 난다면 모듈 다운로드 필요
#cmd에서 pip install 하고 뒤에

from PIL import Image  #pillow
import os
from openpyxl import *   #openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import numpy as np
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import SolidColorFillProperties
from openpyxl.drawing.colors import ColorChoice

#만약 엑셀 파일이 존재하지 않다면 파일 만들 때 사용할 함수
def makeFile():
    wb = Workbook()

    ws = wb.active
    ws.title = "painter"

    saveFile(wb)

#수정한 엑셀 파일을 저장하는 함수
def saveFile(wb : Workbook):
    wb.save(getFilePath("painter.xlsx"))


#엑셀 파일을 불러오는 함수
def openFile():
    wb = load_workbook(getFilePath("painter.xlsx"))
    return wb


#파일들을 쉽게 열 수 있게 도와주는 함수
def getFilePath(f : str):
    path = os.path.dirname(__file__)
    return path + "\\" + f

#이미지를 불러오고 사이즈를 측정
image = Image.open(getFilePath("source.png")).convert("RGB")
size = image.size

#이미지를 numpy로 저장
pix = np.array(image)

#만약 사진을 출력할 엑셀파일이 있다면 원래 시트를 지우고 새로 만들고
#엑셀 파일이 없다면 엑셀파일을 새로 만든다
try:
    wb = openFile()
    wb.create_sheet("pre")
    wb.remove_sheet(wb["painter"])
    wb["pre"].title = "painter"
    wb.remove_sheet(wb["statistics"])
except:
    makeFile()
    wb = openFile()

#사진을 출력할 워크시트를 불러온다
ws = wb.active

#사진에 해당하는 영역의 픽셀 크기를 작은 1x1 정사각형이 되도록 만든다
for i in range(1,size[0]+1):
    ws.column_dimensions[get_column_letter(i)].width = 1

for i in range(1,size[1]+1):
    ws.row_dimensions[i].height = 7

#픽셀의 rgb값 통계를 작성할 딕셔너리 생성
dic = {}

#사진의 픽셀의 rgb값을 구하여 해당하는 엑셀의 셀에 색을 채운다
for i in range(1,size[0]+1):
    for j in range(1,size[1]+1):
        cell = get_column_letter(i)+str(j)
        k = pix[j-1][i-1].tolist()
        rgb = (format(k[0],"02x")+format(k[1],"02x")+format(k[2],"02x")).upper()
        
        #딕셔너리에 픽셀의 rgb값 저장
        if rgb in dic.keys():
            dic[rgb]+=1
        else:
            dic[rgb]=1

        print(rgb)
        ws[cell].fill = PatternFill(start_color=rgb,fill_type="solid")

#저장한 픽셀 자료를 크기순으로 정렬
dic = dict(sorted(dic.items(),key=lambda x: x[1],reverse=True))

#사진의 픽셀 rgb값의 통계를 정리할 시트 생성
wb.create_sheet("statistics")

#새로 만든 시트에 통계를 적을 표 생성
st = wb["statistics"]
stlist = [("rgb 코드","사용 횟수")]

for key in dic.keys():
    stlist.append((key,dic[key]))

for row in stlist:
    st.append(row)

#원 그래프 만들기
pie = PieChart()
pie.style = 39
labels = Reference(st,min_col=1,min_row=2,max_row=len(dic.keys())+2)
data = Reference(st,min_col=2,min_row=2,max_row=len(dic.keys())+2)

pie.add_data(data,titles_from_data=True)
pie.set_categories(labels)
pie.title = "픽셀에 사용된 rpg값"


#가장 많은 색을 튀어나오게하고 해당 색으로 내부 채우기
colors = list(dic.keys())

i=0
slice = DataPoint(
    idx=0,
    explosion=10,
    spPr=GraphicalProperties(solidFill=ColorChoice(srgbClr=colors[i]))
)

pie.series[i].data_points = [slice]
pie.layout
st.add_chart(pie,"E1")

#수정이 왼료된 파일을 저장한다
saveFile(wb)